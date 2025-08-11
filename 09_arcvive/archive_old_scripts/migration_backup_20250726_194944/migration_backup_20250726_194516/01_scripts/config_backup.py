# 🕒 2025-07-01-14-30-00
# SCRPA_LAPTOP/config.py
# Author: R. A. Carucci
# Purpose: Configuration settings for SCRPA report automation - LAPTOP VERSION

import os
import logging
import openpyxl
from datetime import datetime, date, timedelta
from typing import Union, List, Tuple

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Helper Function for Date Conversion ---
def _to_date_object(input_date: Union[str, datetime, date]) -> date:
    """
    Converts various date input types (str, datetime) to a date object.
    Raises ValueError if conversion fails.
    """
    if isinstance(input_date, str):
        try:
            return datetime.strptime(input_date, "%Y_%m_%d").date()
        except ValueError:
            raise ValueError(f"Invalid date string format. Expected YYYY_MM_DD, got: {input_date}")
    elif isinstance(input_date, datetime):
        return input_date.date()
    elif isinstance(input_date, date):
        return input_date
    else:
        raise ValueError(f"Unsupported date format. Expected str, datetime, or date, got: {type(input_date)}")

# --- CONFIGURATION CONSTANTS ---

# Dynamic user home directory for portability
USER_HOME = os.path.expanduser('~')

# LAPTOP-SPECIFIC FILE PATHS
LAPTOP_ROOT = os.path.join(USER_HOME, "SCRPA_LAPTOP")

# Excel files - OneDrive should sync to same location
EXCEL_CYCLE_PATH = os.path.join(USER_HOME, r"OneDrive - City of Hackensack\_Hackensack_Data_Repository\7Day_28Day_Cycle_20250414.xlsx")
EXCEL_CYCLE_WORKSHEET_NAME = "7Day_28Day_250414"

# PowerPoint templates
POWERPOINT_TEMPLATES = {
    "main": os.path.join(USER_HOME, r"OneDrive - City of Hackensack\_25_SCRPA\TIME_Based\TimeSCRPATemplet\SCRPA_REPORT.pptx"),
    "7day": os.path.join(USER_HOME, r"OneDrive - City of Hackensack\_25_SCRPA\TIME_Based\7_Day_Templet_SCRPA_Time\SCRPA_REPORT.pptx")
}

# Report output directory
REPORT_BASE_DIR = os.path.join(USER_HOME, r"OneDrive - City of Hackensack\_25_SCRPA\TIME_Based\Reports")

# ArcGIS Pro project path
APR_PATH = os.path.join(USER_HOME, r"OneDrive - City of Hackensack\_25_SCRPA\TIME_Based\TimeSCRPATemplet\TimeSCRPATemplet.aprx")

# Crime type folder mapping
CRIME_FOLDER_MAPPING = {
    "MV Theft": "MV_Theft",
    "Burglary - Auto": "Burglary_Auto", 
    "Burglary - Comm & Res": "Burglary_Comm_And_Res",
    "Robbery": "Robbery",
    "Sexual Offenses": "Sexual_Offenses"
}

# SQL patterns for crime types
CRIME_SQL_PATTERNS = {
    "MV Theft": ["MOTOR VEHICLE THEFT", "AUTO THEFT"],
    "Burglary - Auto": "BURGLARY FROM MOTOR VEHICLE",
    "Burglary - Comm & Res": ["BURGLARY", "BREAKING & ENTERING"],
    "Robbery": "ROBBERY",
    "Sexual Offenses": ["SEXUAL ASSAULT", "RAPE", "SEXUAL OFFENSE"]
}

# --- EXCEL INTEGRATION FUNCTIONS ---

def get_excel_report_name(target_date):
    """Get report name from Excel based on target date."""
    try:
        if isinstance(target_date, str):
            target_date = datetime.strptime(target_date, "%Y_%m_%d").date()
        elif isinstance(target_date, datetime):
            target_date = target_date.date()
            
        workbook = openpyxl.load_workbook(EXCEL_CYCLE_PATH, data_only=True)
        worksheet = workbook[EXCEL_CYCLE_WORKSHEET_NAME]
        
        # Look for the date in column A (Report_Due_Date)
        for row in range(2, worksheet.max_row + 1):
            cell_value = worksheet[f'A{row}'].value
            if cell_value and isinstance(cell_value, date) and cell_value == target_date:
                report_name = worksheet[f'F{row}'].value  # Report_Name is in column F
                workbook.close()
                return report_name
                
        workbook.close()
        return None
        
    except Exception as e:
        print(f"Excel lookup error: {e}")
        return None

def get_excel_7day_period(target_date):
    """Get 7-day period dates from Excel."""
    try:
        if isinstance(target_date, str):
            target_date = datetime.strptime(target_date, "%Y_%m_%d").date()
        elif isinstance(target_date, datetime):
            target_date = target_date.date()
            
        workbook = openpyxl.load_workbook(EXCEL_CYCLE_PATH, data_only=True)
        worksheet = workbook[EXCEL_CYCLE_WORKSHEET_NAME]
        
        for row in range(2, worksheet.max_row + 1):
            due_date = worksheet[f'A{row}'].value  # Report_Due_Date is in column A
            if due_date and isinstance(due_date, date) and due_date == target_date:
                start_date = worksheet[f'B{row}'].value  # 7_Day_Start is in column B
                end_date = worksheet[f'C{row}'].value    # 7_Day_End is in column C
                workbook.close()
                return start_date, end_date
                
        workbook.close()
        return None, None
        
    except Exception as e:
        print(f"Excel period lookup error: {e}")
        return None, None

def get_correct_folder_name(target_date):
    """Get correct folder name using Excel lookup."""
    if isinstance(target_date, str):
        try:
            target_date = datetime.strptime(target_date, "%Y_%m_%d").date()
            date_str = target_date.strftime("%Y_%m_%d")
        except ValueError:
            date_str = target_date
    elif isinstance(target_date, (date, datetime)):
        if isinstance(target_date, datetime):
            target_date = target_date.date()
        date_str = target_date.strftime("%Y_%m_%d")
    else:
        date_str = str(target_date)
    
    report_name = get_excel_report_name(target_date)
    
    if report_name:
        return f"{report_name}_{date_str}_7Day"
    else:
        # Fallback to old naming if Excel lookup fails
        print("Warning: Using fallback folder naming")
        return f"C05W24_{date_str}_7Day"

def get_7day_period_dates(target_date):
    """Get the correct 7-day period dates from Excel (preferred) or calculation (fallback)"""
    
    # First try to get from Excel
    start_date, end_date = get_excel_7day_period(target_date)
    
    if start_date and end_date:
        print(f"📊 Using Excel 7-day period: {start_date} to {end_date}")
        return start_date, end_date
    
    # Fallback to calculation if Excel lookup fails
    print("⚠️ Excel lookup failed, using calculated period")
    if isinstance(target_date, str):
        target_date = datetime.strptime(target_date, "%Y_%m_%d").date()
    elif isinstance(target_date, datetime):
        target_date = target_date.date()
    
    # For reports due on target_date, we want the 7 days BEFORE the due date
    end_date = target_date - timedelta(days=1)  # Day before report due
    start_date = end_date - timedelta(days=6)   # 7 days total (inclusive)
    
    print(f"📊 Using calculated 7-day period: {start_date} to {end_date}")
    return start_date, end_date

def get_crime_type_folder(crime_type, target_date_str):
    """Get crime-specific output folder path"""
    main_folder = get_correct_folder_name(target_date_str)
    base_path = os.path.join(REPORT_BASE_DIR, main_folder)
    
    # Get the standardized folder name for this crime type
    folder_name = CRIME_FOLDER_MAPPING.get(crime_type, crime_type.replace(' ', '_'))
    
    full_path = os.path.join(base_path, folder_name)
    os.makedirs(full_path, exist_ok=True)
    return full_path

def get_standardized_filename(crime_type):
    """Get standardized filename prefix for crime type"""
    return CRIME_FOLDER_MAPPING.get(crime_type, crime_type.replace(' ', '_'))

def get_output_folder(date_str):
    """Get main output folder path using Excel naming."""
    return os.path.join(REPORT_BASE_DIR, get_correct_folder_name(date_str))

def get_sql_pattern_for_crime(crime_type):
    """Get SQL pattern for specific crime type"""
    return CRIME_SQL_PATTERNS.get(crime_type, crime_type)

# =============================================================================
# VALIDATION FUNCTIONS
# =============================================================================

def validate_all_paths():
    """Validate that all paths in config.py exist and are accessible"""
    print("🔍 VALIDATING CONFIG PATHS")
    print("=" * 40)
    
    paths_to_check = {
        "Excel Cycle File": EXCEL_CYCLE_PATH,
        "Report Base Directory": REPORT_BASE_DIR,
        "ArcGIS Pro Project": APR_PATH,
        "Main PowerPoint Template": POWERPOINT_TEMPLATES.get("main", ""),
        "7-Day PowerPoint Template": POWERPOINT_TEMPLATES.get("7day", "")
    }
    
    all_valid = True
    
    for name, path in paths_to_check.items():
        if path and os.path.exists(path):
            print(f"✅ {name}")
            print(f"   📁 {path}")
        else:
            print(f"❌ {name} NOT FOUND")
            print(f"   📁 {path}")
            all_valid = False
        print()
    
    if all_valid:
        print("🎉 ALL PATHS VALIDATED - READY TO RUN!")
        return True
    else:
        print("\n⚠️ SOME TESTS FAILED - CHECK PATHS")
        return False

def quick_config_test():
    """Quick test of Excel integration"""
    print("🧪 TESTING EXCEL INTEGRATION")
    print("=" * 50)
    
    try:
        # Test date: Today
        test_date = date.today()
        test_date_str = test_date.strftime("%Y_%m_%d")
        
        print(f"📅 Test date: {test_date}")
        print(f"📅 Test date string: {test_date_str}")
        
        # Test Excel lookup
        report_name = get_excel_report_name(test_date)
        print(f"📋 Excel report name: {report_name}")
        
        # Test folder name generation
        folder_name = get_correct_folder_name(test_date)
        print(f"📁 Folder name: {folder_name}")
        
        # Test 7-day period calculation
        start_date, end_date = get_7day_period_dates(test_date)
        print(f"📊 7-Day period: {start_date} to {end_date}")
        
        # Test crime type folder
        test_crime = "MV Theft"
        crime_folder = get_crime_type_folder(test_crime, test_date_str)
        print(f"🚔 Crime folder for {test_crime}: {crime_folder}")
        
        print("=" * 50)
        print("✅ Excel integration test complete!")
        
        if report_name:
            print("🎉 EXCEL INTEGRATION IS WORKING!")
            return True
        else:
            print("⚠️ EXCEL LOOKUP FAILED - CHECK EXCEL DATA")
            return False
            
    except Exception as e:
        print(f"❌ Test failed with error: {e}")
        return False

# =============================================================================
# MAIN EXECUTION
# =============================================================================

if __name__ == "__main__":
    # Run the comprehensive test when this file is executed directly
    quick_config_test()
    print("\n" + "=" * 60)
    validate_all_paths()